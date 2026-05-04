import pandas as pd
import pytest
from openpyxl import load_workbook

from etl.sources import init_db, load_provider_master, load_timesheet
from etl.reconcile import reconcile
from etl.report import write_report
from etl.generator import generate_timesheet, write_timesheet_xlsx


@pytest.fixture
def result(tmp_path):
    db = tmp_path / "master.db"
    init_db(db)
    master = load_provider_master(db)
    ts_df = generate_timesheet(seed=42)
    ts_path = tmp_path / "ts.xlsx"
    write_timesheet_xlsx(ts_df, ts_path)
    ts = load_timesheet(ts_path)
    return reconcile(ts, master)


def test_report_creates_five_sheets(tmp_path, result):
    out = tmp_path / "report.xlsx"
    write_report(result, out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {
        "Summary", "Provider Utilization", "Exceptions", "Department Rollup", "Unknown Providers"
    }


def test_summary_sheet_has_total_providers(tmp_path, result):
    out = tmp_path / "report.xlsx"
    write_report(result, out)
    wb = load_workbook(out)
    ws = wb["Summary"]
    cells = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value) for r in range(1, 12)]
    labels = [c[0] for c in cells if c[0]]
    assert "Total providers" in labels


def test_provider_utilization_sheet_has_rows(tmp_path, result):
    out = tmp_path / "report.xlsx"
    write_report(result, out)
    wb = load_workbook(out)
    ws = wb["Provider Utilization"]
    assert ws.max_row > 1


def test_exceptions_sheet_filters_to_flagged_only(tmp_path, result):
    out = tmp_path / "report.xlsx"
    write_report(result, out)
    df = pd.read_excel(out, sheet_name="Exceptions")
    if not df.empty:
        assert (df["exceptions"] != "").all()
