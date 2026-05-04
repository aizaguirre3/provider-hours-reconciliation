"""Write a formatted multi-sheet Excel report from a reconciliation result."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def write_report(result, output_path):
    """Write Summary, Provider Utilization, Exceptions, Department Rollup, Unknown Providers."""
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_summary(summary_ws, result["summary"])

    util_ws = wb.create_sheet("Provider Utilization")
    _write_table(util_ws, result["utilization"])

    util = result["utilization"]
    exceptions = util[util["exceptions"] != ""].reset_index(drop=True)
    exc_ws = wb.create_sheet("Exceptions")
    _write_table(exc_ws, exceptions)

    dept_ws = wb.create_sheet("Department Rollup")
    _write_table(dept_ws, result["department_rollup"])

    unk_ws = wb.create_sheet("Unknown Providers")
    _write_table(unk_ws, result["unknown_providers"])

    wb.save(output_path)


def _write_summary(ws, summary):
    ws["A1"] = "Provider Hours Reconciliation — Weekly Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    rows = [
        ("Total providers", summary["total_providers"]),
        ("Total contracted hours", summary["total_contracted_hours"]),
        ("Total actual hours", summary["total_actual_hours"]),
        ("Overall utilization", f"{summary['overall_utilization_pct'] * 100:.1f}%"),
        ("Total estimated cost", f"${summary['total_estimated_cost']:,.2f}"),
        ("Providers with exceptions", summary["n_with_exceptions"]),
        ("Unknown providers in timesheet", summary["n_unknown_providers"]),
        ("Timesheet rows dropped (bad data)", summary["dropped_rows"]),
    ]

    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    for cell in (ws["A3"], ws["B3"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, (k, v) in enumerate(rows, start=4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A4"


def _write_table(ws, df):
    if df is None or df.empty:
        ws["A1"] = "(no rows)"
        return
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(ws.columns, start=1):
        values = [str(c.value) for c in col if c.value is not None]
        max_len = max((len(v) for v in values), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 36)
